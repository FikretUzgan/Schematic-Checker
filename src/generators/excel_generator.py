import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

class ExcelGenerator:
    """Generates a styled Excel report for component rating verification."""
    
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.colors = {
            'NOK': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),      # Light Red
            'Marginal': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), # Light Yellow
            'OK': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),        # Light Green
            'AuditFail': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),     # Red
            'AuditWarn': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')      # Light Blue
        }

    def generate(self, results: list):
        """Creates the Excel file with Summary, Details, Library Errors, and Derating Errors sheets."""
        df = pd.DataFrame(results)
        
        # Sort results: NOK -> Marginal -> OK (Handle strings like "NOK (Switching)")
        def get_sort_key(v):
            v_str = str(v)
            if v_str.startswith('NOK'): return 0
            if v_str.startswith('Marginal'): return 1
            if v_str.startswith('OK'): return 2
            return 3
            
        df['sort_order'] = df['Verdict'].apply(get_sort_key)
        df = df.sort_values('sort_order').drop(columns=['sort_order'])
        
        # Filter for Library Errors and Derating Errors
        library_errors = df[df['AuditReason'].str.contains('Library Error', case=False, na=False)]
        derating_errors = df[df['Verdict'].str.startswith('NOK', na=False) | df['Verdict'].str.startswith('Marginal', na=False)]

        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            # 1. Summary Sheet (first for visibility)
            summary_data = self._create_summary_data(df, library_errors, derating_errors)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            workbook = writer.book
            summary_sheet = workbook['Summary']
            self._style_summary_sheet(summary_sheet)
            
            # 2. Details Sheet
            df.to_excel(writer, sheet_name='Verification Details', index=False)
            details_sheet = workbook['Verification Details']
            self._style_details_sheet(details_sheet, df)
            
            # 3. Library Errors Sheet
            if not library_errors.empty:
                library_errors.to_excel(writer, sheet_name='Library Errors', index=False)
                lib_errors_sheet = workbook['Library Errors']
                self._style_details_sheet(lib_errors_sheet, library_errors)
            
            # 4. Derating Errors Sheet
            if not derating_errors.empty:
                derating_errors.to_excel(writer, sheet_name='Derating Errors', index=False)
                derating_sheet = workbook['Derating Errors']
                self._style_details_sheet(derating_sheet, derating_errors)

        print(f"Excel report generated: {os.path.abspath(self.output_path)}")

    def _style_details_sheet(self, sheet, df):
        # Header formatting
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Verdict-based coloring
        # Note: We use the capitalized 'Verdict' key consistent with the rest of the app
        verdict_col_idx = df.columns.get_loc('Verdict') + 1
        for row_idx in range(2, sheet.max_row + 1):
            verdict_val = str(sheet.cell(row=row_idx, column=verdict_col_idx).value)
            fill = None
            if verdict_val.startswith('NOK'): fill = self.colors['NOK']
            elif verdict_val.startswith('Marginal'): fill = self.colors['Marginal']
            elif verdict_val.startswith('OK'): fill = self.colors['OK']
            elif verdict_val.startswith('Unknown'): fill = self.colors['NOK']  # Treat Unknown as NOK
            
            if fill:
                for col_idx in range(1, len(df.columns) + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = fill

        # Audit-based coloring (overrides verdict color if critical)
        audit_col_idx = df.columns.get_loc('AuditVerdict') + 1 if 'AuditVerdict' in df.columns else None
        if audit_col_idx:
            for row_idx in range(2, sheet.max_row + 1):
                audit_val = sheet.cell(row=row_idx, column=audit_col_idx).value
                if audit_val == 'FAIL':
                    fill = self.colors['AuditFail']
                    for col_idx in range(1, len(df.columns) + 1):
                        sheet.cell(row=row_idx, column=col_idx).fill = fill
                elif audit_val == 'WARNING' and not str(sheet.cell(row=row_idx, column=verdict_col_idx).value).startswith('NOK'):
                    # Only color for warning if not already red from a verdict
                    sheet.cell(row=row_idx, column=audit_col_idx).fill = self.colors['AuditWarn']

        # Auto-adjust column width
        for i, col in enumerate(df.columns, 1):
            if col == 'Reason' or col == 'AuditReason':
                width = 30
            else:
                width = 15
            sheet.column_dimensions[get_column_letter(i)].width = width

    def _create_summary_data(self, df, library_errors, derating_errors):
        if df.empty: return []
        
        counts = df['Type'].value_counts().to_dict()
        verdicts = df['Verdict'].value_counts().to_dict()
        
        summary = [
            {'Category': 'Total Components Checked', 'Value': len(df)},
            {'Category': 'Verdict: OK', 'Value': len(df[df['Verdict'].str.startswith('OK', na=False)])},
            {'Category': 'Verdict: Marginal', 'Value': len(df[df['Verdict'].str.startswith('Marginal', na=False)])},
            {'Category': 'Verdict: NOK', 'Value': len(df[df['Verdict'].str.startswith('NOK', na=False)])},
            {'Category': '-- Errors Summary --', 'Value': ''},
            {'Category': 'Library Errors', 'Value': len(library_errors)},
            {'Category': 'Derating Errors', 'Value': len(derating_errors)},
            {'Category': '-- Breakdown by Type --', 'Value': ''}
        ]
        
        for t, count in counts.items():
            summary.append({'Category': f"Type: {t}", 'Value': count})
            
        return summary

    def _style_summary_sheet(self, sheet):
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 15
        for cell in sheet['A']:
            if cell.value and ('Verdict:' in str(cell.value) or 'Total' in str(cell.value)):
                cell.font = Font(bold=True)
